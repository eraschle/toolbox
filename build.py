# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Skillsliste-Generator
=====================

Liest die Inhalte aus  skills_daten.xlsx  und erzeugt daraus die fertige
Webseite  Skillsliste.html  (auf Basis von  template.html).

Aufruf:  uv run build.py      (oder einfach build.bat doppelklicken)

Es wird KEIN Code in der HTML-Datei verändert – nur die Inhalte (Texte und
Emoji) werden aus der Excel-Datei eingesetzt. Das Layout/Design steckt
unverändert in template.html.
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

# Konsole auf UTF-8 stellen, sonst scheitert das Drucken von Emoji/Umlauten.
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "skills_daten.xlsx"
TEMPLATE = ROOT / "template.html"
OUTPUT = ROOT / "Skillsliste.html"
PLACEHOLDER = "var DATA = /*__BUILD_DATA__*/{};"

# Anzeige-Name (Excel)  ->  interner Schlüssel (HTML/JS, darf sich NICHT ändern)
STUFE_KEY = {"Hoch": "hoch", "Mittel": "mittel", "Tief": "tief"}
STUFE_ORDER = ["hoch", "mittel", "tief"]


class BuildError(Exception):
    """Sammelt eine oder mehrere verständliche Fehlermeldungen."""

    def __init__(self, messages):
        self.messages = messages if isinstance(messages, list) else [messages]
        super().__init__("\n".join(self.messages))


# ── Hilfsfunktionen ──────────────────────────────────────────
def clean(value) -> str:
    """Zellwert -> getrimmter String. Leere Zelle -> '' (nie 'None'/'nan')."""
    if value is None:
        return ""
    return str(value).strip()


def slug(label: str) -> str:
    """Erzeugt eine interne, stabile ID aus dem Kategorie-Label.

    Nur intern verwendet (aktiver Tab-Zustand im JS) – nie sichtbar.
    """
    text = label.lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    text = text.replace("ß", "ss")
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "kat"


def get_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise BuildError(
            f"Das Blatt '{name}' fehlt in skills_daten.xlsx. "
            f"Vorhandene Blätter: {', '.join(wb.sheetnames)}"
        )
    return wb[name]


def read_rows(ws, expected_header):
    """Liest ein Blatt als Liste von dicts {Spaltenname: Wert}.

    Liefert zusätzlich die echte Excel-Zeilennummer für Fehlermeldungen.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BuildError(f"Das Blatt '{ws.title}' ist leer.")
    header = [clean(h) for h in rows[0]]
    missing = [h for h in expected_header if h not in header]
    if missing:
        raise BuildError(
            f"Im Blatt '{ws.title}' fehlen die Spalten: {', '.join(missing)}. "
            f"Bitte die Kopfzeile nicht umbenennen."
        )
    idx = {name: header.index(name) for name in expected_header}
    out = []
    for excel_row, raw in enumerate(rows[1:], start=2):
        record = {name: clean(raw[i]) if i < len(raw) else "" for name, i in idx.items()}
        if not any(record.values()):
            continue  # komplett leere Zeile überspringen
        record["_row"] = excel_row
        out.append(record)
    return out


# ── Daten lesen & prüfen ─────────────────────────────────────
def load_data():
    if not XLSX.exists():
        raise BuildError(f"Datei nicht gefunden: {XLSX.name}")

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    stufen_rows = read_rows(
        get_sheet(wb, "Stufen"),
        ["Stufe", "Bezeichnung", "Bereich", "Icon", "Intro", "Farbe", "Farbe2", "Hell"],
    )
    kat_rows = read_rows(get_sheet(wb, "Kategorien"), ["Stufe", "Kategorie", "Icon"])
    skill_rows = read_rows(
        get_sheet(wb, "Skills"),
        ["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp"],
    )
    wb.close()

    errors = []

    def check_stufe(rec, blatt):
        if rec["Stufe"] not in STUFE_KEY:
            errors.append(
                f"Blatt '{blatt}', Zeile {rec['_row']}: Stufe '{rec['Stufe']}' "
                f"ist ungültig – erlaubt sind nur Hoch, Mittel oder Tief."
            )
            return None
        return STUFE_KEY[rec["Stufe"]]

    # Stufen-Metadaten
    stufen = {}
    for rec in stufen_rows:
        key = check_stufe(rec, "Stufen")
        if key:
            stufen[key] = rec
    fehlend = [s for s in STUFE_ORDER if s not in stufen]
    if fehlend:
        namen = ", ".join(k for k, v in STUFE_KEY.items() if v in fehlend)
        errors.append(f"Im Blatt 'Stufen' fehlen Zeilen für: {namen}.")

    # Kategorien (Reihenfolge je Stufe + Icon)
    kat_order = {s: [] for s in STUFE_ORDER}   # key -> [(label, icon)]
    kat_icon = {}                              # (key, label) -> icon
    for rec in kat_rows:
        key = check_stufe(rec, "Kategorien")
        if not key:
            continue
        label = rec["Kategorie"]
        if not label:
            errors.append(f"Blatt 'Kategorien', Zeile {rec['_row']}: Kategorie fehlt.")
            continue
        if (key, label) in kat_icon:
            errors.append(
                f"Blatt 'Kategorien', Zeile {rec['_row']}: Kategorie '{label}' "
                f"ist in Stufe '{rec['Stufe']}' doppelt."
            )
            continue
        kat_icon[(key, label)] = rec["Icon"]
        kat_order[key].append((label, rec["Icon"]))

    # Skills den Kategorien zuordnen (Reihenfolge bleibt erhalten)
    skills_by = {}  # (key, label) -> [skill-dict]
    for rec in skill_rows:
        key = check_stufe(rec, "Skills")
        if not key:
            continue
        label = rec["Kategorie"]
        for feld in ("Emoji", "Titel", "Beschreibung"):
            if not rec[feld]:
                errors.append(
                    f"Blatt 'Skills', Zeile {rec['_row']}: '{feld}' darf nicht leer sein."
                )
        if (key, label) not in kat_icon:
            errors.append(
                f"Blatt 'Skills', Zeile {rec['_row']}: Kategorie '{label}' "
                f"existiert in Stufe '{rec['Stufe']}' nicht "
                f"(bitte zuerst im Blatt 'Kategorien' anlegen)."
            )
            continue
        skills_by.setdefault((key, label), []).append(
            {"e": rec["Emoji"], "t": rec["Titel"], "b": rec["Beschreibung"], "tip": rec["Tipp"]}
        )

    if errors:
        raise BuildError(errors)

    # DATA-Struktur im erwarteten Schema bauen
    data = {}
    for key in STUFE_ORDER:
        meta = stufen[key]
        kategorien = []
        for label, icon in kat_order[key]:
            kategorien.append(
                {
                    "id": slug(label),
                    "label": label,
                    "icon": icon,
                    "skills": skills_by.get((key, label), []),
                }
            )
        data[key] = {
            "label": meta["Bezeichnung"],
            "bereich": meta["Bereich"],
            "farbe": meta["Farbe"],
            "farbe2": meta["Farbe2"],
            "hell": meta["Hell"],
            "icon": meta["Icon"],
            "intro": meta["Intro"],
            "kategorien": kategorien,
        }
    return data


# ── HTML schreiben ───────────────────────────────────────────
def render(data: dict):
    if not TEMPLATE.exists():
        raise BuildError(f"Vorlage nicht gefunden: {TEMPLATE.name}")
    template = TEMPLATE.read_bytes().decode("utf-8-sig")  # evtl. BOM entfernen
    if PLACEHOLDER not in template:
        raise BuildError(
            f"Platzhalter nicht in {TEMPLATE.name} gefunden. "
            f"Die Vorlage muss '{PLACEHOLDER}' enthalten."
        )
    payload = json.dumps(data, ensure_ascii=False, separators=(", ", ": "))
    html = template.replace(PLACEHOLDER, f"var DATA = {payload};", 1)
    # gleiche Datei-Konvention wie das Original: UTF-8 mit BOM, CRLF
    OUTPUT.write_bytes(b"\xef\xbb\xbf" + html.encode("utf-8"))


def main():
    try:
        data = load_data()
        render(data)
    except BuildError as exc:
        print("\n❌ Build abgebrochen – bitte folgendes in skills_daten.xlsx korrigieren:\n")
        for msg in exc.messages:
            print(f"   • {msg}")
        print()
        sys.exit(1)

    total = sum(len(k["skills"]) for d in data.values() for k in d["kategorien"])
    print("✅ Skillsliste.html wurde neu erstellt.")
    print(f"   Stufen: {len(data)} | Kategorien: "
          f"{sum(len(d['kategorien']) for d in data.values())} | Skills: {total}")


if __name__ == "__main__":
    main()
