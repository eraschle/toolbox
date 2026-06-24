# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Excel neu aufsetzen (nur für Entwickler / Erst-Einrichtung)
===========================================================

ACHTUNG: Dieses Skript ÜBERSCHREIBT skills_daten.xlsx komplett mit den
Inhalten aus der aktuellen Skillsliste.html. Es dient nur dem erstmaligen
Erzeugen bzw. dem Zurücksetzen der Excel-Datei – NICHT der laufenden Pflege.
Für die normale Pflege wird skills_daten.xlsx direkt bearbeitet.

Aufruf:  uv run tools/seed_excel.py

Besonderheiten der erzeugten Datei:
- Tipp-Spalte OHNE führende Glühbirne (💡 wird beim Build automatisch ergänzt).
- AutoFilter (Filter-Pfeile) in der Kopfzeile aller Blätter.
- Emoji-/Icon-Spalten in Schriftart "Segoe UI Emoji" für farbige Darstellung.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
HTML = ROOT / "docs" / "Skillsliste.html"
OUT = ROOT / "skills_daten.xlsx"

STUFE_DISPLAY = {"hoch": "Hoch", "mittel": "Mittel", "tief": "Tief"}

HEADER_FILL = PatternFill("solid", fgColor="E65100")
HEADER_FONT = Font(bold=True, color="FFFFFF")
EMOJI_FONT = Font(name="Segoe UI Emoji", size=12)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def load_data() -> dict:
    text = HTML.read_bytes().decode("utf-8-sig")
    for line in text.replace("\r\n", "\n").split("\n"):
        if line.lstrip().startswith("var DATA"):
            return json.loads(line[line.index("{") : line.rindex("}") + 1])
    raise SystemExit("var DATA nicht in Skillsliste.html gefunden")


def strip_birne(tip: str) -> str:
    """Entfernt eine führende Glühbirne 💡 (+ folgende Leerzeichen)."""
    tip = (tip or "").strip()
    if tip.startswith("💡"):
        tip = tip[1:].lstrip()
    return tip


def style_sheet(ws, ncols, widths, emoji_col=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions          # Filter-Pfeile in Kopfzeile
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if emoji_col:
        for row in ws.iter_rows(min_row=2, min_col=emoji_col, max_col=emoji_col):
            row[0].font = EMOJI_FONT
            row[0].alignment = CENTER


def add_stufe_dropdown(ws, col_letter):
    dv = DataValidation(type="list", formula1='"Hoch,Mittel,Tief"', allow_blank=False)
    dv.error = "Bitte Hoch, Mittel oder Tief wählen."
    dv.errorTitle = "Ungültige Stufe"
    dv.prompt = "Hoch / Mittel / Tief"
    dv.promptTitle = "Stufe"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max(ws.max_row, 2)}")


def main():
    data = load_data()
    wb = openpyxl.Workbook()

    # ── Blatt Skills ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Skills"
    ws.append(["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp"])
    for key, d in data.items():
        for kat in d["kategorien"]:
            for s in kat["skills"]:
                ws.append(
                    [
                        STUFE_DISPLAY[key],
                        kat["label"],
                        s.get("e", ""),
                        s.get("t", ""),
                        s.get("b", ""),
                        strip_birne(s.get("tip", "")),
                    ]
                )
    style_sheet(ws, 6, [9, 18, 8, 26, 60, 55], emoji_col=3)
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = WRAP_TOP   # Beschreibung
        row[5].alignment = WRAP_TOP   # Tipp
    add_stufe_dropdown(ws, "A")

    # ── Blatt Stufen ─────────────────────────────────────────
    ws2 = wb.create_sheet("Stufen")
    ws2.append(["Stufe", "Bezeichnung", "Bereich", "Icon", "Intro",
                "Farbe", "Farbe2", "Hell"])
    for key, d in data.items():
        ws2.append([STUFE_DISPLAY[key], d.get("label", ""), d.get("bereich", ""),
                    d.get("icon", ""), d.get("intro", ""), d.get("farbe", ""),
                    d.get("farbe2", ""), d.get("hell", "")])
    style_sheet(ws2, 8, [9, 22, 16, 8, 60, 11, 11, 11], emoji_col=4)
    for row in ws2.iter_rows(min_row=2):
        row[4].alignment = WRAP_TOP   # Intro
    add_stufe_dropdown(ws2, "A")

    # ── Blatt Kategorien ─────────────────────────────────────
    ws3 = wb.create_sheet("Kategorien")
    ws3.append(["Stufe", "Kategorie", "Icon"])
    for key, d in data.items():
        for kat in d["kategorien"]:
            ws3.append([STUFE_DISPLAY[key], kat["label"], kat.get("icon", "")])
    style_sheet(ws3, 3, [9, 20, 8], emoji_col=3)
    add_stufe_dropdown(ws3, "A")

    wb.save(OUT)
    print(f"Geschrieben: {OUT}")
    print(f"  Skills: {ws.max_row - 1} | Stufen: {ws2.max_row - 1} | "
          f"Kategorien: {ws3.max_row - 1}")


if __name__ == "__main__":
    main()
